import os
import sys
from openpyxl import load_workbook
import re
from openpyxl.utils import get_column_letter

# 常量定义
SHEET_NAME = '评审信息'
ARCHIVE_ADDRESS_FIELD = '评审文件归档地址'
REVIEWER_FIELD = '评审人员'
REVIEWER_TO_ADD = '吴建良'
FILE_KEYWORD = '评审报告'
EXCEL_EXTENSION = '.xlsx' # 假定为现代 Excel 文件格式

def modify_review_report_file(file_path):
    """
    修改指定的 Excel 文件，根据llm/gc_review-report_modify.md中的要求进行处理。
    如果文件被成功修改并保存，则返回 True，否则返回 False。
    """
    print(f"尝试处理文件: {file_path}")
    try:
        workbook = load_workbook(filename=file_path)
        print(f"  - 成功打开文件。")
    except Exception as e:
        print(f"  - 警告: 无法打开或读取文件 '{file_path}': {e}")
        return False

    if SHEET_NAME not in workbook.sheetnames:
        print(f"  - 警告: 文件 '{file_path}' 中不包含名为 '{SHEET_NAME}' 的工作表。")
        return False
    print(f"  - 成功找到工作表 '{SHEET_NAME}'。")

    sheet = workbook[SHEET_NAME]
    file_modified = False

    # 存储字段名单元格的坐标 (行, 列)
    archive_address_field_cell_coord = None
    reviewer_field_cell_coord = None

    print(f"  - 正在全表查找字段: '{ARCHIVE_ADDRESS_FIELD}' 和 '{REVIEWER_FIELD}'")

    # 遍历所有单元格来查找字段名
    # openpyxl 的 max_row 和 max_column 可能会返回比实际数据行/列更大的值
    # 实际迭代时，使用已知的最大行/列（或设置一个合理上限，例如前50行）
    # 为了保险起见，遍历整个used range
    for row_idx in range(1, sheet.max_row + 1):
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell_value = str(cell.value if cell.value is not None else '').strip()
            # print(f"    - 检查单元格 ({row_idx}, {col_idx}): '{cell_value}'") # 过于详细，暂时注释

            if cell_value == ARCHIVE_ADDRESS_FIELD:
                archive_address_field_cell_coord = (row_idx, col_idx)
                print(f"    - 找到 '{ARCHIVE_ADDRESS_FIELD}' 在单元格 ({row_idx}, {get_column_letter(col_idx)})")
            elif cell_value == REVIEWER_FIELD:
                reviewer_field_cell_coord = (row_idx, col_idx)
                print(f"    - 找到 '{REVIEWER_FIELD}' 在单元格 ({row_idx}, {get_column_letter(col_idx)})")

            # 如果两个字段都已找到，则停止搜索以提高效率
            if archive_address_field_cell_coord and reviewer_field_cell_coord:
                break
        if archive_address_field_cell_coord and reviewer_field_cell_coord:
            break

    if not archive_address_field_cell_coord and not reviewer_field_cell_coord:
        print(f"  - 警告: 未在文件 '{file_path}' 的 '{SHEET_NAME}' 工作表中找到 '{ARCHIVE_ADDRESS_FIELD}' 和 '{REVIEWER_FIELD}' 字段。跳过此文件。")
        return False

    # 处理 '评审人员' 字段
    if reviewer_field_cell_coord:
        field_row, field_col = reviewer_field_cell_coord
        # 假设值在字段名单元格的右侧相邻单元格
        reviewer_value_cell = sheet.cell(row=field_row, column=field_col + 1)
        current_reviewers_raw = str(reviewer_value_cell.value if reviewer_value_cell.value is not None else '').strip()
        print(f"  - '评审人员' 的值在单元格 ({field_row}, {get_column_letter(field_col+1)})。当前值: '{current_reviewers_raw}'")

        # 统一分隔符：将所有指定的中文和英文逗号、分号替换为统一的中文顿号
        # 并移除可能存在的连续顿号，以及字符串首尾的顿号
        normalized_reviewers = re.sub(r'[,;，；]', '、', current_reviewers_raw)
        normalized_reviewers = re.sub(r'、+', '、', normalized_reviewers).strip('、')

        # 将评审人员拆分为列表，去除空项和多余空格
        reviewer_list = [name.strip() for name in normalized_reviewers.split('、') if name.strip()]

        # 检查是否需要添加 REVIEWER_TO_ADD
        added_new_reviewer = False
        if REVIEWER_TO_ADD not in reviewer_list:
            reviewer_list.append(REVIEWER_TO_ADD)
            added_new_reviewer = True

        # 重新组合成最终的规范化字符串
        final_reviewers_string = '、'.join(reviewer_list)
        
        # 检查是否需要更新单元格的值
        if current_reviewers_raw != final_reviewers_string or added_new_reviewer:
            reviewer_value_cell.value = final_reviewers_string
            file_modified = True
            print(f"  - '评审人员' 字段已更新为: '{reviewer_value_cell.value}'。")
        else:
            print(f"  - '评审人员' 已包含 '{REVIEWER_TO_ADD}' 且分隔符已规范化，无需修改。")
    else:
        print(f"  - 未找到 '{REVIEWER_FIELD}' 字段，跳过相关处理。")

    #rows_to_delete = []
    ## 处理 '评审文件归档地址' 字段
    #if archive_address_field_cell_coord:
    #    field_row, field_col = archive_address_field_cell_coord
    #    # 假设值在字段名单元格的右侧相邻单元格
    #    archive_address_value_cell = sheet.cell(row=field_row, column=field_col + 1)
    #    archive_address_value = str(archive_address_value_cell.value if archive_address_value_cell.value is not None else '').strip()
    #    print(f"  - '{ARCHIVE_ADDRESS_FIELD}' 的值在单元格 ({field_row}, {get_column_letter(field_col+1)})。当前值: '{archive_address_value}'")
    #
    #    if archive_address_value == '':
    #        rows_to_delete.append(field_row)
    #        file_modified = True
    #        print(f"  - 行 {field_row} 被标记为删除，因为 '{ARCHIVE_ADDRESS_FIELD}' 的值为空。")
    #    else:
    #        print(f"  - '{ARCHIVE_ADDRESS_FIELD}' 的值不为空，不删除行 {field_row}。")
    #else:
    #    print(f"  - 未找到 '{ARCHIVE_ADDRESS_FIELD}' 字段，跳过相关处理。")

    ## 执行行删除操作 (从最高行索引到最低行索引，确保删除正确)
    #if rows_to_delete:
    #    print(f"  - 准备删除行: {sorted(rows_to_delete, reverse=True)}")
    #    for row_idx_to_del in sorted(rows_to_delete, reverse=True):
    #        sheet.delete_rows(row_idx_to_del, 1)
    #        print(f"  - 成功删除行: {row_idx_to_del}")
    #else:
    #    print(f"  - 没有行需要删除。")

    if file_modified:
        try:
            workbook.save(filename=file_path)
            print(f"成功修改文件并保存: {file_path}")
            return True
        except Exception as e:
            print(f"错误: 无法保存文件 {file_path}: {e}。请检查文件是否被其他程序占用或权限不足。")
            return False
    else:
        print(f"文件 '{file_path}' 未进行任何修改。")
    return False

def main():
    if len(sys.argv) < 2:
        print("用法: python gc_review-report-modify.py <目录路径>")
        sys.exit(1)

    target_directory = sys.argv[1]

    if not os.path.isdir(target_directory):
        print(f"错误: 指定的路径不是一个有效的目录: {target_directory}")
        sys.exit(1)

    found_files_count = 0
    for root, _, files in os.walk(target_directory):
        for filename in files:
            # 检查文件名是否包含关键词并且是 Excel 文件
            if FILE_KEYWORD in filename and filename.endswith(EXCEL_EXTENSION):
                file_path = os.path.join(root, filename)
                print(f"DEBUG: 发现并尝试处理文件: {file_path}") # 临时调试输出
                found_files_count += 1
                modify_review_report_file(file_path)

    if found_files_count == 0:
        print(f"INFO: 在 '{target_directory}' 及其子目录中未找到任何名为 '{FILE_KEYWORD}' 或以 '{EXCEL_EXTENSION}' 结尾的文件。") # 临时调试输出

if __name__ == "__main__":
    main()
