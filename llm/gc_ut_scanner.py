import os
import argparse
import csv
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

def get_compiler_from_excel(filepath):
    """
    从Excel文件中名为'封面页'的sheet中获取'拟制'字段后的值作为编制人。
    """
    try:
        workbook = load_workbook(filepath, read_only=True)
        if '封面页' not in workbook.sheetnames:
            return ""

        sheet = workbook['封面页']
        compiler = ""
        # 遍历所有单元格查找“拟制”
        for row in sheet.iter_rows():
            for i, cell in enumerate(row):
                if cell.value == "拟制":
                    # 如果找到“拟制”，则其右侧的单元格是编制人
                    if i + 1 < len(row):
                        compiler_cell = row[i+1]
                        if compiler_cell.value:
                            compiler = str(compiler_cell.value).strip()
                            return compiler
        return ""
    except InvalidFileException:
        # 文件不是有效的Excel文件
        return ""
    except Exception as e:
        # 捕获其他可能的错误，例如文件损坏或权限问题
        # print(f"Error reading {filepath}: {e}") # for debug
        return ""

def scan_and_print_unit_tests(scan_path):
    """
    扫描指定路径下的单元测试文件并打印CSV格式的汇总信息。
    """
    results = []
    for root, _, files in os.walk(scan_path):
        for filename in files:
            # 过滤文件名中包含“单元测试”且为excel表格的文件
            if "单元测试" in filename and filename.endswith(".xlsx"):
                filepath = os.path.join(root, filename)

                # 获取编制日期并转换格式
                try:
                    parts = filename.split('-')
                    date_str = parts[-1].split('.')[0] # 获取最后一个字段，并移除文件扩展名
                    formatted_date = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:]}"
                except IndexError:
                    formatted_date = "" # 无法解析日期
                
                # 获取编制人
                compiler = get_compiler_from_excel(filepath)

                # 获取文档路径（父目录的目录名）
                parent_dir_name = os.path.basename(os.path.dirname(filepath))

                results.append({
                    "文档类型": "单元测试报告",
                    "文档名称": filename,
                    "编制人": compiler,
                    "文档路径": parent_dir_name,
                    "编制日期": formatted_date
                })
    
    # 打印CSV头部
    print("序号,文档类型,文档名称,编制人,文档路径,编制日期")
    # 打印CSV数据
    for i, item in enumerate(results):
        print(f"{i + 1},{item['文档类型']},{item['文档名称']},{item['编制人']},{item['文档路径']},{item['编制日期']}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="GC单元测试文档扫描工具")
    parser.add_argument("scan_path", help="待扫描的目录路径")
    args = parser.parse_args()

    scan_and_print_unit_tests(args.scan_path)
